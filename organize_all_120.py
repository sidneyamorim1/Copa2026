import openpyxl
import json
import csv
import unicodedata

excel_path = "/Users/sidneyamorim/Desktop/Ajuste/Grupos Final_120.xlsx"
seed_path = "/Users/sidneyamorim/Desktop/Ajuste/src/seed.json"
output_all_path = "/Users/sidneyamorim/Desktop/Ajuste/palpites_corrigidos.csv"
output_aline_path = "/Users/sidneyamorim/Desktop/Ajuste/palpites_aline.csv"

# Map to correct/official team names in Portuguese
team_mapping = {
    'Africa': 'África do Sul',
    'Arab Saudita': 'Arábia Saudita',
    'Arábia': 'Arábia Saudita',
    'Austria': 'Áustria',
    'Aústria': 'Áustria',
    'Cabo': 'Cabo Verde',
    'Cabo verde': 'Cabo Verde',
    'Cabo Verde': 'Cabo Verde',
    'Corea': 'Coreia do Sul',
    'Coréa': 'Coreia do Sul',
    'Cost Marfim': 'Costa do Marfim',
    'Haití': 'Haiti',
    'Jordania': 'Jordânia',
    'Noroega': 'Noruega',
    'Nova Zelandia': 'Nova Zelândia',
    'Tunisia': 'Tunísia',
    'Uzbequistão': 'Uzbequistão',
    'Belgica': 'Bélgica',
    'portugal': 'Portugal',
    'croácia': 'Croácia',
    'colombia': 'Colômbia',
    'gana': 'Gana',
    'equador': 'Equador',
    'Congo': 'RD Congo',  # Map to 'RD Congo' to match seed.json
    'Árgélia': 'Argélia',
    'Inglaterra ': 'Inglaterra',
}

def clean_team_name(name):
    if not name:
        return ""
    name_str = str(name).strip()
    return team_mapping.get(name_str, name_str)

def clean_score(val):
    if val is None or val == "":
        return ""
    return str(val).strip()

def normalize(name):
    if not name:
        return ""
    n = str(name).strip().lower()
    n = "".join(c for c in unicodedata.normalize("NFD", n) if unicodedata.category(c) != "Mn")
    
    # manual normalization mappings
    if n in ('corea', 'coreia', 'corea do sul'): return 'coreia do sul'
    if n == 'bosnia': return 'bosnia e herz.'
    if n in ('costa de marfim', 'cost marfim'): return 'costa do marfim'
    if n in ('catar', 'qatar'): return 'catar'
    if n == 'panama': return 'panama'
    if n == 'tunisia': return 'tunisia'
    if n == 'nova zelandia': return 'nova zelandia'
    if n in ('arabia', 'arab saudita'): return 'arabia saudita'
    if n == 'austria': return 'austria'
    if n == 'australia': return 'australia'
    if n == 'belgica': return 'belgica'
    if n == 'portugal': return 'portugal'
    if n == 'croacia': return 'croacia'
    if n == 'colombia': return 'colombia'
    if n == 'gana': return 'gana'
    if n == 'equador': return 'equador'
    if n == 'congo': return 'rd congo'
    if n == 'rd congo': return 'rd congo'
    if n == 'argelia': return 'argelia'
    if n == 'inglaterra': return 'inglaterra'
    
    return n

# Load seed games to know the database home/away team orders
with open(seed_path, 'r', encoding='utf-8') as f:
    seed_games = json.load(f)

db_games = {}
for g in seed_games:
    hc = normalize(g['time_casa'])
    fa = normalize(g['time_fora'])
    key = tuple(sorted([hc, fa]))
    db_games[key] = (g['time_casa'], g['time_fora'])

def get_db_match_and_order(sheet_home, sheet_away):
    h_clean = clean_team_name(sheet_home)
    a_clean = clean_team_name(sheet_away)
    h_norm = normalize(h_clean)
    a_norm = normalize(a_clean)
    key = tuple(sorted([h_norm, a_norm]))
    
    if key in db_games:
        db_home, db_away = db_games[key]
        if normalize(db_home) == h_norm:
            return db_home, db_away, False
        else:
            return db_home, db_away, True
    else:
        # Fallback to cleaned sheet names
        return h_clean, a_clean, False

valid_participants = {'sidney', 'eduardo', 'aline', 'matheus', 'mateus', 'silvio', 'daniel'}

wb = openpyxl.load_workbook(excel_path, data_only=True)

all_csv_rows = []
aline_predictions = []

# ----------------- PARSE 'Grupos' SHEET -----------------
sheet_grupos = wb['Grupos']
block_starts_1based = [2, 14, 26, 38, 50, 62, 74, 85, 97, 109, 121, 133]

matchday_data_grupos = []
max_row_g = sheet_grupos.max_row

for r in range(1, max_row_g + 1):
    val_b = sheet_grupos.cell(row=r, column=2).value
    val_c = sheet_grupos.cell(row=r, column=3).value
    val_d = sheet_grupos.cell(row=r, column=4).value
    
    # Header row detection
    if val_b == 'Data' and val_c == 'Hr' and val_d == 'Nome':
        current_teams = {}
        current_actuals = {}
        for j, start_col in enumerate(block_starts_1based):
            home_team = sheet_grupos.cell(row=r, column=start_col + 4).value
            home_goals = sheet_grupos.cell(row=r, column=start_col + 5).value
            away_goals = sheet_grupos.cell(row=r, column=start_col + 7).value
            away_team = sheet_grupos.cell(row=r, column=start_col + 8).value
            if home_team and away_team and home_team != 'Gols' and away_team != 'Gols':
                current_teams[j] = (home_team, away_team)
                current_actuals[j] = (clean_score(home_goals), clean_score(away_goals))
        matchday_data_grupos.append({
            'teams': current_teams,
            'actuals': current_actuals,
            'predictions': []
        })
        continue
    
    # Prediction row detection
    if val_d and str(val_d).strip().lower() in valid_participants:
        if len(matchday_data_grupos) > 0:
            row_preds = {}
            for j, start_col in enumerate(block_starts_1based):
                if j not in matchday_data_grupos[-1]['teams']:
                    continue
                p_name = sheet_grupos.cell(row=r, column=start_col + 2).value
                name_to_use = str(p_name).strip() if p_name else str(val_d).strip()
                home_goals = sheet_grupos.cell(row=r, column=start_col + 5).value
                away_goals = sheet_grupos.cell(row=r, column=start_col + 7).value
                row_preds[j] = (name_to_use, clean_score(home_goals), clean_score(away_goals))
            matchday_data_grupos[-1]['predictions'].append(row_preds)

# Process and output Group stage rows
for md in matchday_data_grupos:
    teams_dict = md['teams']
    actuals_dict = md['actuals']
    preds_list = md['predictions']
    
    for j in sorted(teams_dict.keys()):
        sheet_h, sheet_a = teams_dict[j]
        act_h, act_a = actuals_dict[j]
        
        # Get correct database order
        db_home, db_away, is_reversed = get_db_match_and_order(sheet_h, sheet_a)
        
        # Swap goals if reversed
        final_act_h = act_a if is_reversed else act_h
        final_act_a = act_h if is_reversed else act_a
        
        # Oficial row
        all_csv_rows.append({
            'Nome': 'Oficial',
            'Time Casa': db_home,
            'Gols Casa': final_act_h,
            'Time Fora': db_away,
            'Gols Fora': final_act_a
        })
        
        # User predictions
        for user_pred in preds_list:
            if j in user_pred:
                username, pred_h, pred_a = user_pred[j]
                if username.lower() == 'mateus':
                    username = 'Matheus'
                
                final_pred_h = pred_a if is_reversed else pred_h
                final_pred_a = pred_h if is_reversed else pred_a
                
                pred_dict = {
                    'Nome': username,
                    'Time Casa': db_home,
                    'Gols Casa': final_pred_h,
                    'Time Fora': db_away,
                    'Gols Fora': final_pred_a
                }
                all_csv_rows.append(pred_dict)
                if username.lower() == 'aline':
                    aline_predictions.append(pred_dict)

print(f"Parsed 'Grupos': {len(all_csv_rows)} rows generated.")

# ----------------- PARSE '16 avos' SHEET -----------------
sheet_16 = wb['16 avos']
max_row_16 = sheet_16.max_row
count_16 = 0

for r in range(1, max_row_16 + 1):
    val_b = sheet_16.cell(row=r, column=2).value
    val_c = sheet_16.cell(row=r, column=3).value
    val_d = sheet_16.cell(row=r, column=4).value
    val_e = sheet_16.cell(row=r, column=5).value
    
    # Header row detection
    if val_b == 'Data' and val_c == 'Hr' and val_d == 'Nome' and val_e == 'Gols':
        home_team = sheet_16.cell(row=r, column=6).value
        home_goals = sheet_16.cell(row=r, column=7).value
        away_goals = sheet_16.cell(row=r, column=9).value
        away_team = sheet_16.cell(row=r, column=10).value
        
        if home_team and away_team:
            db_home, db_away, is_reversed = get_db_match_and_order(home_team, away_team)
            
            act_h = clean_score(home_goals)
            act_a = clean_score(away_goals)
            
            final_act_h = act_a if is_reversed else act_h
            final_act_a = act_h if is_reversed else act_a
            
            # Add Oficial row for 16-avos
            all_csv_rows.append({
                'Nome': 'Oficial',
                'Time Casa': db_home,
                'Gols Casa': final_act_h,
                'Time Fora': db_away,
                'Gols Fora': final_act_a
            })
            count_16 += 1
            
            # Read predictions in subsequent rows
            offset = 1
            while r + offset <= max_row_16:
                p_row = r + offset
                p_name = sheet_16.cell(row=p_row, column=4).value
                if p_name and str(p_name).strip().lower() in valid_participants:
                    username = str(p_name).strip()
                    if username.lower() == 'mateus':
                        username = 'Matheus'
                        
                    pred_h = sheet_16.cell(row=p_row, column=7).value
                    pred_a = sheet_16.cell(row=p_row, column=9).value
                    
                    final_pred_h = clean_score(pred_a) if is_reversed else clean_score(pred_h)
                    final_pred_a = clean_score(pred_h) if is_reversed else clean_score(pred_a)
                    
                    pred_dict = {
                        'Nome': username,
                        'Time Casa': db_home,
                        'Gols Casa': final_pred_h,
                        'Time Fora': db_away,
                        'Gols Fora': final_pred_a
                    }
                    all_csv_rows.append(pred_dict)
                    if username.lower() == 'aline':
                        aline_predictions.append(pred_dict)
                    offset += 1
                else:
                    break

print(f"Parsed '16 avos': {count_16} games found. Total rows generated: {len(all_csv_rows)}")

# Write all predictions to palpites_corrigidos.csv using comma
with open(output_all_path, 'w', encoding='utf-8', newline='') as f:
    writer = csv.DictWriter(f, fieldnames=['Nome', 'Time Casa', 'Gols Casa', 'Time Fora', 'Gols Fora'], delimiter=',')
    writer.writeheader()
    for pred in all_csv_rows:
        writer.writerow(pred)

# Write Aline's predictions
with open(output_aline_path, 'w', encoding='utf-8', newline='') as f:
    writer = csv.DictWriter(f, fieldnames=['Nome', 'Time Casa', 'Gols Casa', 'Time Fora', 'Gols Fora'], delimiter=';')
    writer.writeheader()
    for pred in aline_predictions:
        writer.writerow(pred)

print("Processamento completo!")
